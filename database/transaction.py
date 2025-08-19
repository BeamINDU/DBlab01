from database.connect_to_db import engine, Session, text, SQLAlchemyError
from fastapi import HTTPException
import database.schemas as schemas
from datetime import datetime
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class TransactionDB:
    def _fetch_one(self, query: str, params: dict):
        try:
            with engine.connect() as conn:
              result = conn.execute(text(query), params)
              return result.mappings().first()
        except SQLAlchemyError as e:
            print(f"Database error: {e}")
            return []

    def _fetch_all(self, query: str, params: dict = None):
        try:
            with engine.connect() as conn:
                if params:
                    result = conn.execute(text(query), params)
                else:
                    result = conn.execute(text(query))
                return list(result.mappings())
        except SQLAlchemyError as e:
            print(f"Database error: {e}")
            return []

    def get_transaction(self, model: schemas.TransactionSearch, pagination: bool = True):
        where_filters = []
        params = {}

        # --- Allowed fields for sorting ---
        allowed_order_fields = [ "actualstartdatetime", "actualenddatetime", "prodlot", "prodid", "prodname", "quantity" ]
        order_by = model.order_by if model.order_by in allowed_order_fields else "actualstartdatetime"
        order_dir = "DESC" if str(model.order_dir).lower() == "desc" else "ASC"

        # --- Helper for adding filters ---
        def add_filter(condition, key, value):
            if value is not None:
                where_filters.append(condition)
                params[key] = value

        add_filter("t.actualstartdatetime >= :startdate", "startdate", model.startdate)
        add_filter("t.actualenddatetime <= :enddate", "enddate", model.enddate)
        add_filter("t.prodlot ILIKE :prodlot", "prodlot", f"%{model.prodlot}%" if model.prodlot else None)
        add_filter("t.prodid ILIKE :prodid", "prodid", f"%{model.prodid}%" if model.prodid else None)
        add_filter("p.prodname ILIKE :prodname", "prodname", f"%{model.prodname}%" if model.prodname else None)

        where_clause = f"WHERE {' AND '.join(where_filters)}" if where_filters else ""

        # --- Pagination ---
        limit_clause = ""
        if pagination:
            page = max(model.page or 1, 1)
            page_size = min(max(model.pageSize or 10, 1), 100)
            offset = (page - 1) * page_size
            limit_clause = "LIMIT :limit OFFSET :offset"
            params["limit"] = page_size
            params["offset"] = offset

        # --- Base Select ---
        base_query = """
            FROM transactionreport t
            LEFT JOIN product p ON p.prodid = t.prodid
        """

        # --- Main Query ---
        main_query = f"""
            SELECT
                t.transactionid, t.prodlot, t.prodid,
                p.prodname, t.actualstartdatetime,
                t.actualenddatetime, t.quantity
            {base_query}
            {where_clause}
            ORDER BY {order_by} {order_dir}
            {limit_clause}
        """

        # --- Count Query ---
        count_query = f"""
            SELECT COUNT(*) AS count
            FROM (
                SELECT 1
                {base_query}
                {where_clause}
            ) AS subquery
        """

        # --- Execute ---
        total = self._fetch_one(count_query, params)["count"]
        items = self._fetch_all(main_query, params)

        return {
            "total": total,
            "items": items
        }

    def export_transaction(self, model: schemas.TransactionSearch):
        result = self.get_transaction(model, pagination=False)
        items = result["items"]

        df = pd.DataFrame(items, columns=[
            "actualstartdatetime", "actualenddatetime", "prodlot",
            "prodid", "prodname", "quantity"
        ])

        # เพิ่มคอลัมน์ลำดับ
        df.insert(0, "No.", range(1, len(df) + 1))

        # เปลี่ยนชื่อคอลัมน์ให้สวยงาม
        df.columns = [
            "No.", "Start Date", "End Date", "Lot No",
            "Product ID", "Product Name", "Actual Total Quantity"
        ]

        for col in ["Start Date", "End Date"]:
          if col in df.columns:
              df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')

        output = io.BytesIO()

        if model.export_type and model.export_type.lower() == "csv":
            df.to_csv(output, index=False)
            output.seek(0)
            media_type = "text/csv"
            filename = "transaction.csv"
        else:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sheet_name = "Transaction"
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # เพิ่มความกว้างของคอลัมน์ตามความยาวข้อความ
                for i, column in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(i)
                    max_length = max(df[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[column_letter].width = max_length + 4

                # กรอบเซลล์แบบบาง
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                # สีพื้นหลังของหัวตาราง
                header_fill = PatternFill(start_color='FFBCE0FD', end_color='FFBCE0FD', fill_type='solid')

                # ฟอนต์หัวตาราง: ตัวหนา สีดำ
                header_font = Font(bold=True, color='FF000000')

                # การจัดตำแหน่ง
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                align_right = Alignment(horizontal='right', vertical='center')

                # Style header row
                for cell in worksheet[1]:
                  cell.fill = header_fill
                  cell.font = header_font
                  cell.border = border
                  cell.alignment = align_center

                # Style data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                  for cell in row:
                      cell.border = border
                      if isinstance(cell.value, (int, float)):
                          cell.alignment = align_right  # ตัวเลข ชิดขวา
                      else:
                          cell.alignment = align_left  # ข้อความ ชิดซ้าย
                
            output.seek(0)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "transaction.xlsx"

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


    def suggest_transaction_lotno(self, q: str):
        rows = self._fetch_all("""
            SELECT DISTINCT prodlot FROM transactionreport
            WHERE LOWER(prodlot) LIKE LOWER(:keyword)
            ORDER BY prodlot ASC
            LIMIT 10; """,
            {"keyword": q + "%"}
        )
        return [{"value": row["prodlot"], "label": row["prodlot"]} for row in rows]
