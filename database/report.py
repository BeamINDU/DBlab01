from database.connect_to_db import engine, Session, text, SQLAlchemyError
from fastapi import HTTPException
import database.schemas as schemas
from fastapi.responses import JSONResponse
import database.schemas as schemas
from typing import Union, Dict, Any
from sqlalchemy import text 
from datetime import datetime, date
# from database.images import ImagesService
from fastapi.responses import StreamingResponse
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def error_response(code: int, message: str):
    return JSONResponse( status_code=code, content={"detail": {"error": message}} )

def success_response(code: int, content: Union[Dict[str, Any], str]):
    return JSONResponse( status_code=code, content=content)

class ReportDB:
    def _fetch_one(self, query: str, params: dict):
        try:
            with engine.connect() as conn:
              result = conn.execute(text(query), params)
              return result.mappings().first()
        except SQLAlchemyError as e:
            print(f"Database error: {e}")
            return []

    def _fetch_all(self, query: str, params: dict = None):
        try:
            with engine.connect() as conn:
                if params:
                    result = conn.execute(text(query), params)
                else:
                    result = conn.execute(text(query))
                return list(result.mappings())
        except SQLAlchemyError as e:
            print(f"Database error: {e}")
            return []
    
    #--- Report Product Defect Result -------------------------------------------------------------
    
    def get_report_product_defect(self, model: schemas.ReportProductSearch, pagination: bool = True):
        where_filters = []
        having_filters = []
        params = {}

        allowed_order_fields = [ "defecttime", "prodid", "prodname", "prodseq", "cameraid", "cameraname", "prodstatus", "defectdetail" ]
        order_by = model.order_by if model.order_by in allowed_order_fields else "defecttime"
        order_dir = "DESC" if str(model.order_dir).lower() == "desc" else "ASC"

        # --- Helper for filters ---
        def add_where(condition, key, value):
            if value is not None:
                where_filters.append(condition)
                params[key] = value

        def add_having(condition, key, value):
            if value is not None:
                having_filters.append(condition)
                params[key] = value

        # --- WHERE filters ---
        add_where("pr.defecttime >= :startdate", "startdate", model.startdate)
        add_where("pr.defecttime <= :enddate", "enddate", model.enddate)
        add_where("pr.prodid ILIKE :prodid", "prodid", f"%{model.prodid}%" if model.prodid else None)
        add_where("p.prodname ILIKE :prodname", "prodname", f"%{model.prodname}%" if model.prodname else None)
        add_where("pr.cameraid ILIKE :cameraid", "cameraid", f"%{model.cameraid}%" if model.cameraid else None)
        add_where("c.cameraname ILIKE :cameraname", "cameraname", f"%{model.cameraname}%" if model.cameraname else None)

        # --- HAVING filters ---
        add_having("CASE WHEN BOOL_OR(pr.prodstatus = 'NG') THEN 'NG' ELSE 'OK' END = :prodstatus", "prodstatus", model.prodstatus)
        add_having("STRING_AGG(d.defecttype, ', ' ORDER BY d.defecttype) ILIKE :defecttype", "defecttype", f"%{model.defecttype}%" if model.defecttype else None)

        where_clause = f"WHERE {' AND '.join(where_filters)}" if where_filters else ""
        having_clause = f"HAVING {' AND '.join(having_filters)}" if having_filters else ""

        # --- Pagination ---
        limit_clause = ""
        if pagination:
            page = max(model.page or 1, 1)
            page_size = min(max(model.pageSize or 10, 1), 100)
            offset = (page - 1) * page_size
            limit_clause = "LIMIT :limit OFFSET :offset"
            params["limit"] = page_size
            params["offset"] = offset

        group_by_clause = """
            GROUP BY 
                pr.cameraid, c.cameraname, pr.prodid, p.prodname,
                pr.prodseq, pr.defecttime, pr.imagepath
        """

        base_select = f"""
            FROM productdefectresult pr
            LEFT JOIN product p ON p.prodid = pr.prodid
            LEFT JOIN camera c ON c.cameraid = pr.cameraid
            LEFT JOIN defecttype d ON d.defectid = pr.defectid
            {where_clause}
            {group_by_clause}
            {having_clause}
        """

        # --- Main Query ---
        main_query = f"""
            SELECT *
            FROM (
                SELECT
                    ROW_NUMBER() OVER (ORDER BY pr.defecttime) AS runningno,
                    pr.defecttime,
                    pr.prodid,
                    p.prodname,
                    pr.prodseq,
                    pr.cameraid,
                    c.cameraname,
                    pr.imagepath,
                    CASE WHEN BOOL_OR(pr.prodstatus = 'NG') THEN 'NG' ELSE 'OK' END AS prodstatus,
                    STRING_AGG(d.defecttype, ', ' ORDER BY d.defecttype) AS defectdetail,
                    STRING_AGG(pr.comment, ', ' ORDER BY pr.resultid) AS comment
                {base_select}
            ) AS subquery
            ORDER BY {order_by} {order_dir}
            {limit_clause}
        """

        # --- Count Query ---
        count_query = f"""
            SELECT COUNT(*) AS count
            FROM (
                SELECT 1
                {base_select}
            ) AS count
        """

        total = self._fetch_one(count_query, params)["count"]
        items = self._fetch_all(main_query, params)

        return {
            "total": total,
            "items": items
        }

    def export_report_product_defect(self, model: schemas.ReportProductSearch):
        result = self.get_report_product_defect(model, pagination=False)
        items = result["items"]

        df = pd.DataFrame(items, columns=[
            "defecttime", "prodid", "prodname",
            "defectdetail", "cameraid", "cameraname", "prodstatus"
        ])

        # เพิ่มลำดับ No.
        df.insert(0, "No.", range(1, len(df) + 1))

        # ตั้งชื่อหัวตารางให้อ่านง่าย
        df.columns = [
            "No.", "Datetime", "Product ID", "Product Name",
            "Defect Detail", "Camera ID", "Camera Name", "Status"
        ]

        for col in ["Datetime"]:
          if col in df.columns:
              df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')

        output = io.BytesIO()

        # Export as CSV
        if model.export_type and model.export_type.lower() == "csv":
            df.to_csv(output, index=False)
            output.seek(0)
            media_type = "text/csv"
            filename = "report_product_defect.csv"
        else:
            # Export as Excel + apply style
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sheet_name = "Product Defect"
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # เพิ่มความกว้างของคอลัมน์ตามความยาวข้อความ
                for i, column in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(i)
                    max_length = max(df[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[column_letter].width = max_length + 4

                # กรอบเซลล์แบบบาง
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                # สีพื้นหลังของหัวตาราง
                header_fill = PatternFill(start_color='FFBCE0FD', end_color='FFBCE0FD', fill_type='solid')

                # ฟอนต์หัวตาราง: ตัวหนา สีดำ
                header_font = Font(bold=True, color='FF000000')

                # การจัดตำแหน่ง
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                align_right = Alignment(horizontal='right', vertical='center')

                # Style header row
                for cell in worksheet[1]:
                  cell.fill = header_fill
                  cell.font = header_font
                  cell.border = border
                  cell.alignment = align_center

                # Style data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                  for cell in row:
                      cell.border = border
                      if isinstance(cell.value, (int, float)):
                          cell.alignment = align_right  # ตัวเลข ชิดขวา
                      else:
                          cell.alignment = align_left  # ข้อความ ชิดซ้าย

            output.seek(0)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "report_product_defect.xlsx"

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    def report_product_defect_detail(self, model: schemas.ReportProductSearchDetail, db: Session) -> Dict[str, Any]:
        try:
            query = text("""
                SELECT
                  pr.defecttime,
                  pr.prodid,
                  p.prodname,
                  p.prodserial,
                  pr.prodseq,
                  pr.cameraid,
                  pr.imagepath,
                  CASE
                      WHEN BOOL_OR(pr.prodstatus = 'NG') THEN 'NG'
                      ELSE 'OK'
                  END AS status,
                  CONCAT_WS(E'\n',
                      CONCAT('OK - ', STRING_AGG(pr.defectid, ', ' ORDER BY pr.resultid) FILTER (WHERE pr.prodstatus = 'OK')),
                      CONCAT('NG - ', STRING_AGG(pr.defectid, ', ' ORDER BY pr.resultid) FILTER (WHERE pr.prodstatus = 'NG'))
                  ) AS defect_summary,
                  STRING_AGG(pr.comment, ', ' ORDER BY pr.resultid) AS comment
              FROM productdefectresult pr
              LEFT JOIN product p ON p.prodid = pr.prodid
              WHERE pr.prodid = :prodid
                AND pr.prodseq = :prodseq
                AND pr.cameraid = :cameraid
                AND pr.imagepath = :imagepath
              GROUP BY pr.cameraid, pr.prodid, p.prodname, p.prodserial, pr.prodseq, pr.defecttime, pr.imagepath
            """)

            params = {
                "prodid": model.prodid,
                "prodseq": model.prodseq,
                "cameraid": model.cameraid,
                "imagepath": model.imagepath
            }

            # print("query",query)
            # print("params",params)

            defect_result = db.execute(query, params).mappings().fetchone()
            if not defect_result:
                return {"error": "No data found"}

            data = dict(defect_result)
            # print("DEBUG - defect_result full:", dict(defect_result))

            history_query = text("""
                SELECT
                    actiondate,
                    status,
                    "comment",
                    actionby
                FROM productdefectresult_log
                WHERE prodid = :prodid
                  AND prodseq = :prodseq
                  AND cameraid = :cameraid
                  AND imagepath = :imagepath
                ORDER BY actiondate DESC, status, "comment"
            """)

            history_params = {
                "prodid": model.prodid,
                "prodseq": model.prodseq,
                "cameraid": model.cameraid,
                "imagepath": model.imagepath
            }

            history_result = db.execute(history_query, history_params).mappings().all()
            data["history"] = [dict(row) for row in history_result]

            # try:
            #     imagedata = ImagesService().result_image(model.imagepath)
            #     data["image64"] = imagedata.results.image_b64
            # except Exception:
            #     data["image64"] = None
            

            return data

        except Exception as e:
            return {"error": str(e)}
    
    def update_product_defect_detail(self, item: schemas.ProductDetailUpdate, db: Session):
        try:
            current_query = text("""
                SELECT comment, prodstatus FROM productdefectresult
                WHERE prodid = :prodid
                  AND prodseq = :prodseq
                  AND cameraid = :cameraid
                  AND imagepath = :imagepath
                  AND DATE_TRUNC('second', defecttime) = TIMESTAMP :defecttime
            """)

            current = db.execute(current_query, {
                "prodid": item.prodid,
                "prodseq": item.prodseq,
                "cameraid": item.cameraid,
                "imagepath": item.imagepath,
                "defecttime": item.defecttime
            }).mappings().first()

            if not current:
                return error_response(404, "Record not found")

            # ถ้าไม่มีการเปลี่ยนแปลงค่า comment หรือ prodstatus → ไม่ต้อง update หรือ insert log
            current_comment = current["comment"] or ""
            new_comment = item.comment or ""

            if current_comment == new_comment and current["prodstatus"] == item.prodstatus:
                return success_response(200, {
                    "result": "No changes detected",
                    "prodid": item.prodid,
                    "prodseq": item.prodseq,
                    "cameraid": item.cameraid,
                    "imagepath": item.imagepath,
                    "datetime": item.defecttime,
                })

            # Update
            update_query = text("""
                UPDATE productdefectresult 
                SET comment = :comment, prodstatus = :prodstatus
                WHERE prodid = :prodid
                  AND prodseq = :prodseq
                  AND cameraid = :cameraid
                  AND imagepath = :imagepath
                  AND DATE_TRUNC('second', defecttime) = TIMESTAMP :defecttime
            """)

            update_params = {
                "comment": item.comment,
                "prodstatus": item.prodstatus,
                "prodid": item.prodid,
                "prodseq": item.prodseq,
                "cameraid": item.cameraid,
                "imagepath": item.imagepath,
                "defecttime": item.defecttime
            }

            result = db.execute(update_query, update_params)

            if result.rowcount == 0:
                db.rollback()
                return error_response(404, "No record updated. Please check the input values.")

            # Log only if values changed
            insert_query = text("""
                INSERT INTO productdefectresult_log 
                (prodid, prodseq, cameraid, imagepath, comment, status, actiondate, actionby)
                VALUES (:prodid, :prodseq, :cameraid, :imagepath, :comment, :status, :actiondate, :actionby)
            """)

            log_params = {
                "prodid": item.prodid,
                "prodseq": item.prodseq,
                "cameraid": item.cameraid,
                "imagepath": item.imagepath,
                "comment": item.comment,
                "status": item.prodstatus,
                "actiondate": datetime.now(),
                "actionby": item.actionby
            }

            db.execute(insert_query, log_params)
            db.commit()

            return success_response(200, {
                "result": "ProductDefectResult updated and logged",
                "prodid": item.prodid,
                "prodseq": item.prodseq,
                "cameraid": item.cameraid,
                "imagepath": item.imagepath,
                "datetime": item.defecttime,
            })

        except SQLAlchemyError as e:
            db.rollback()
            return error_response(500, f"Database error: {str(e)}")

   
    #--- Report Defect Summary -------------------------------------------------------------
    
    def get_defect_summary(self, model: schemas.ReportDefectSearch, pagination: bool = True):
        where_filters = []
        params = {}

        allowed_order_fields = ["prodlot", "prodid", "prodname", "defectid", "defecttype", "totalprod", "totalok", "totalng"]
        order_by = model.order_by if model.order_by in allowed_order_fields else "prodlot"
        order_dir = "DESC" if str(model.order_dir).lower() == "desc" else "ASC"

        def add_filter(condition, key, value):
            if value is not None:
                where_filters.append(condition)
                params[key] = value

        add_filter("ds.prodlot ILIKE :prodlot", "prodlot", f"%{model.prodlot}%" if model.prodlot else None)
        add_filter("ds.prodid ILIKE :prodid", "prodid", f"%{model.prodid}%" if model.prodid else None)
        add_filter("p.prodname ILIKE :prodname", "prodname", f"%{model.prodname}%" if model.prodname else None)
        add_filter("ds.defectid ILIKE :defectid", "defectid", f"%{model.defectid}%" if model.defectid else None)
        add_filter("d.defecttype ILIKE :defecttype", "defecttype", f"%{model.defecttype}%" if model.defecttype else None)

        where_clause = f"WHERE {' AND '.join(where_filters)}" if where_filters else ""

        base_from = """
            FROM defectsummary ds
            LEFT JOIN product p ON p.prodid = ds.prodid
            LEFT JOIN defecttype d ON d.defectid = ds.defectid
        """

        # Pagination
        limit_clause = ""
        if pagination:
            page = max(model.page or 1, 1)
            page_size = min(max(model.pageSize or 10, 1), 100)
            offset = (page - 1) * page_size
            limit_clause = "LIMIT :limit OFFSET :offset"
            params["limit"] = page_size
            params["offset"] = offset

        main_query = f"""
            SELECT
                ds.summaryid,
                ds.prodlot,
                ds.prodid,
                p.prodname,
                ds.defectid,
                d.defecttype,
                ds.totalprod,
                ds.totalok,
                ds.totalng
            {base_from}
            {where_clause}
            ORDER BY {order_by} {order_dir}
            {limit_clause}
        """

        count = 0
        if pagination:
            count_query = f"""
                SELECT COUNT(*) AS count
                FROM (
                    SELECT 1
                    {base_from}
                    {where_clause}
                ) AS count
            """
            count = self._fetch_one(count_query, params)["count"]

        items = self._fetch_all(main_query, params)

        return {
            "total": count if pagination else len(items),
            "items": items
        }

    def export_defect_summary(self, model: schemas.ReportDefectSearch):
        result = self.get_defect_summary(model, pagination=False)
        items = result["items"]

        df = pd.DataFrame(items, columns=[
            "prodlot", "prodid", "prodname", "defectid", "defecttype",
            "totalprod", "totalok", "totalng"
        ])

        # เพิ่มคอลัมน์ No. เริ่มจาก 1
        df.insert(0, "No.", range(1, len(df) + 1))

        df.columns = [
            "No.", "Lot No", "Product ID", "Product Name", "Defect Type ID",
            "Defect Type Name", "Total", "OK %", "NG %"
        ]

        output = io.BytesIO()

        if model.export_type and model.export_type.lower() == "csv":
            df.to_csv(output, index=False)
            output.seek(0)
            media_type = "text/csv"
            filename = "report_defect_summary.csv"
        else:
            # default to Excel
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sheet_name = 'Defect Summary'
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                worksheet = writer.sheets[sheet_name]

                # เพิ่มความกว้างของคอลัมน์ตามความยาวข้อความ
                for i, column in enumerate(df.columns, start=1):
                    column_letter = get_column_letter(i)
                    max_length = max(df[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[column_letter].width = max_length + 4

                # กรอบเซลล์แบบบาง
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                # สีพื้นหลังของหัวตาราง
                header_fill = PatternFill(start_color='FFBCE0FD', end_color='FFBCE0FD', fill_type='solid')

                # ฟอนต์หัวตาราง: ตัวหนา สีดำ
                header_font = Font(bold=True, color='FF000000')

                # การจัดตำแหน่ง
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center')
                align_right = Alignment(horizontal='right', vertical='center')

                # Style header row
                for cell in worksheet[1]:
                  cell.fill = header_fill
                  cell.font = header_font
                  cell.border = border
                  cell.alignment = align_center

                # Style data rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                  for cell in row:
                      cell.border = border
                      if isinstance(cell.value, (int, float)):
                          cell.alignment = align_right  # ตัวเลข ชิดขวา
                      else:
                          cell.alignment = align_left  # ข้อความ ชิดซ้าย

            output.seek(0)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "report_defect_summary.xlsx"

        return StreamingResponse(
            output,
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    def suggest_defect_lotno(self, q: str):
        rows = self._fetch_all("""
            SELECT DISTINCT prodlot FROM defectsummary
            WHERE LOWER(prodlot) LIKE LOWER(:keyword)
            ORDER BY prodlot ASC
            LIMIT 10; """,
            {"keyword": q + "%"}
        )
        return [{"value": row["prodlot"], "label": row["prodlot"]} for row in rows]
    
